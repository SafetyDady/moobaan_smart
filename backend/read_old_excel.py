import openpyxl
import sys

full_path = r'C:\web_project\moobaan_smart\backend\old_journal.xlsx'
print(f'Opening: {full_path}')

wb = openpyxl.load_workbook(full_path, data_only=True)
print(f'Sheet names: {wb.sheetnames}')

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\n{"="*60}')
    print(f'Sheet: {sn} (rows={ws.max_row}, cols={ws.max_column})')
    print(f'{"="*60}')
    
    # Print all rows (or up to 200)
    max_rows = min(ws.max_row + 1, 200)
    for r in range(1, max_rows):
        row_data = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val is not None:
                s = str(val)
                if len(s) > 50:
                    s = s[:50] + '...'
                row_data.append(s)
            else:
                row_data.append('')
        # Skip completely empty rows
        if any(v for v in row_data):
            print(f'  R{r:3d}: {row_data}')
