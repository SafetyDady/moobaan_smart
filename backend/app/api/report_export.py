"""
Phase 5.2: Report Export API (PDF/Excel)

Purpose:
- Export admin data tables as PDF or Excel files
- Supports: invoices, payins, houses, members, expenses
- READ ONLY — no data mutation

Endpoints:
- GET /api/reports/export/{report_type}?format=pdf|xlsx
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, desc
from typing import Optional
from datetime import datetime
import io
import logging

from app.db.session import get_db
from app.core.deps import require_admin_or_accounting
from app.db.models import (
    User, Invoice, InvoiceStatus, PayinReport, PayinStatus,
    House, HouseStatus, Expense, ExpenseStatus
)
from app.db.models.house_member import HouseMember

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports/export", tags=["Report Export"])


# ─── Helper: Format Thai Baht ──────────────────────────────────────────

def fmt_baht(val):
    """Format number as Thai Baht string"""
    if val is None:
        return "฿0"
    return f"฿{val:,.2f}"


def fmt_date(dt):
    """Format datetime to Thai-friendly string"""
    if not dt:
        return "-"
    if isinstance(dt, str):
        return dt
    return dt.strftime("%d/%m/%Y")


def fmt_datetime(dt):
    """Format datetime with time"""
    if not dt:
        return "-"
    if isinstance(dt, str):
        return dt
    return dt.strftime("%d/%m/%Y %H:%M")


# ─── Data Fetchers ──────────────────────────────────────────────────────

def fetch_invoices(db: Session, period: Optional[str] = None):
    """Fetch invoice data for export"""
    query = db.query(Invoice)
    if period:
        query = query.filter(Invoice.period == period)
    query = query.order_by(desc(Invoice.created_at))
    invoices = query.all()

    headers = ["เลขที่", "บ้าน", "งวด", "จำนวนเงิน", "สถานะ", "วันที่สร้าง"]
    rows = []
    for inv in invoices:
        house_code = inv.house.house_code if inv.house else "-"
        rows.append([
            str(inv.id),
            house_code,
            inv.period or "-",
            fmt_baht(inv.amount),
            inv.status.value if inv.status else "-",
            fmt_date(inv.created_at),
        ])
    return headers, rows, f"invoices_{period or 'all'}"


def fetch_payins(db: Session, status_filter: Optional[str] = None):
    """Fetch payin data for export"""
    query = db.query(PayinReport)
    if status_filter:
        try:
            status_enum = PayinStatus(status_filter)
            query = query.filter(PayinReport.status == status_enum)
        except ValueError:
            pass
    query = query.order_by(desc(PayinReport.created_at))
    payins = query.all()

    headers = ["เลขที่", "บ้าน", "จำนวนเงิน", "วันที่โอน", "สถานะ", "วันที่แจ้ง"]
    rows = []
    for p in payins:
        house_code = p.house.house_code if p.house else "-"
        rows.append([
            str(p.id),
            house_code,
            fmt_baht(p.amount),
            fmt_date(p.transfer_date),
            p.status.value if p.status else "-",
            fmt_datetime(p.created_at),
        ])
    return headers, rows, f"payins_{status_filter or 'all'}"


def fetch_houses(db: Session):
    """Fetch house data for export"""
    houses = db.query(House).order_by(House.house_code).all()

    headers = ["รหัสบ้าน", "ที่อยู่", "สถานะ", "ค่าส่วนกลาง/เดือน"]
    rows = []
    for h in houses:
        rows.append([
            h.house_code or "-",
            h.address or "-",
            h.status.value if h.status else "-",
            fmt_baht(h.monthly_fee) if hasattr(h, 'monthly_fee') and h.monthly_fee else "-",
        ])
    return headers, rows, "houses"


def fetch_members(db: Session):
    """Fetch member data for export"""
    members = db.query(HouseMember).order_by(HouseMember.id).all()

    headers = ["ชื่อ-นามสกุล", "บ้าน", "โทรศัพท์", "อีเมล", "บทบาท"]
    rows = []
    for m in members:
        house_code = m.house.house_code if m.house else "-"
        rows.append([
            m.full_name or "-",
            house_code,
            m.phone or "-",
            m.email or "-",
            m.role or "-",
        ])
    return headers, rows, "members"


def fetch_expenses(db: Session, period: Optional[str] = None):
    """Fetch expense data for export"""
    query = db.query(Expense)
    if period:
        query = query.filter(Expense.period == period)
    query = query.order_by(desc(Expense.created_at))
    expenses = query.all()

    headers = ["เลขที่", "รายการ", "หมวดหมู่", "จำนวนเงิน", "สถานะ", "วันที่"]
    rows = []
    for e in expenses:
        rows.append([
            str(e.id),
            e.description or "-",
            e.category.value if hasattr(e, 'category') and e.category else "-",
            fmt_baht(e.amount),
            e.status.value if e.status else "-",
            fmt_date(e.expense_date if hasattr(e, 'expense_date') else e.created_at),
        ])
    return headers, rows, f"expenses_{period or 'all'}"


# ─── PDF Generator ──────────────────────────────────────────────────────

def generate_pdf(title: str, headers: list, rows: list) -> io.BytesIO:
    """Generate PDF report using reportlab"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm,
    )

    elements = []
    styles = getSampleStyleSheet()

    # Try to register Thai font
    thai_font_registered = False
    font_paths = [
        "/usr/share/fonts/truetype/thai/Sarabun-Regular.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansThai-Regular.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    ]
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('ThaiFont', fp))
                thai_font_registered = True
                break
            except Exception:
                continue

    font_name = 'ThaiFont' if thai_font_registered else 'Helvetica'

    # Title
    title_style = styles['Title']
    if thai_font_registered:
        title_style.fontName = font_name
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 5*mm))

    # Subtitle with date
    subtitle_style = styles['Normal']
    if thai_font_registered:
        subtitle_style.fontName = font_name
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    elements.append(Paragraph(f"Generated: {now_str} | Total: {len(rows)} records", subtitle_style))
    elements.append(Spacer(1, 5*mm))

    # Table data
    table_data = [headers] + rows

    # Calculate column widths
    available_width = landscape(A4)[0] - 30*mm
    col_count = len(headers)
    col_width = available_width / col_count

    table = Table(table_data, colWidths=[col_width] * col_count)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        # Body
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        # Alternating rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    # Set Thai font for all cells if available
    if thai_font_registered:
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
        ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


# ─── Excel Generator ────────────────────────────────────────────────────

def generate_excel(title: str, headers: list, rows: list) -> io.BytesIO:
    """Generate Excel report using openpyxl"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Date row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total: {len(rows)} records")
    date_cell.font = Font(size=10, color="666666")
    date_cell.alignment = Alignment(horizontal="center")

    # Headers (row 4)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for row_idx, row_data in enumerate(rows, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if (row_idx - 5) % 2 == 1:
                cell.fill = alt_fill

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row in rows:
            if col_idx - 1 < len(row):
                max_length = max(max_length, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = min(max_length + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ─── Export Endpoint ────────────────────────────────────────────────────

REPORT_TYPES = {
    "invoices": {"title": "รายงานใบแจ้งหนี้", "fetcher": fetch_invoices},
    "payins": {"title": "รายงานการชำระเงิน", "fetcher": fetch_payins},
    "houses": {"title": "รายงานบ้านทั้งหมด", "fetcher": fetch_houses},
    "members": {"title": "รายงานสมาชิก", "fetcher": fetch_members},
    "expenses": {"title": "รายงานค่าใช้จ่าย", "fetcher": fetch_expenses},
}


@router.get("/{report_type}")
async def export_report(
    report_type: str,
    format: str = Query("xlsx", regex="^(pdf|xlsx)$", description="Export format: pdf or xlsx"),
    period: Optional[str] = Query(None, description="Filter by period (YYYY-MM)"),
    status: Optional[str] = Query(None, description="Filter by status"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_accounting),
):
    """
    Export report data as PDF or Excel file.
    
    Supported report_type: invoices, payins, houses, members, expenses
    """
    if report_type not in REPORT_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid report type. Supported: {', '.join(REPORT_TYPES.keys())}"
        )

    config = REPORT_TYPES[report_type]
    title = config["title"]

    # Fetch data based on report type
    try:
        if report_type == "invoices":
            headers, rows, filename = config["fetcher"](db, period=period)
        elif report_type == "payins":
            headers, rows, filename = config["fetcher"](db, status_filter=status)
        elif report_type == "expenses":
            headers, rows, filename = config["fetcher"](db, period=period)
        else:
            headers, rows, filename = config["fetcher"](db)
    except Exception as e:
        logger.error(f"Export data fetch failed: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch report data")

    # Generate file
    try:
        if format == "pdf":
            buffer = generate_pdf(title, headers, rows)
            media_type = "application/pdf"
            ext = "pdf"
        else:
            buffer = generate_excel(title, headers, rows)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ext = "xlsx"
    except Exception as e:
        logger.error(f"Export generation failed: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate report file")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    download_filename = f"{filename}_{timestamp}.{ext}"

    logger.info(f"📊 Report exported: {download_filename} by user {current_user.id}")

    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{download_filename}"'
        }
    )
