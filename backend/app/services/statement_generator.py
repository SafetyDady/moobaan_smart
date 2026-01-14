"""
Statement generation services for PDF and Excel export.

Provides production-ready statement generation with:
- Bilingual Thai/English content
- Official formatting and disclaimers
- Thai font support for PDF
- Accounting-friendly Excel format
"""

import io
import os
import hashlib
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, List, Any

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import Settings


def _get_short_hash(text: str) -> str:
    """Generate short hash for document reference."""
    return hashlib.md5(text.encode()).hexdigest()[:8].upper()


def _format_currency(amount: Decimal) -> str:
    """Format currency with Thai comma formatting."""
    return f"{amount:,.2f}"


def _convert_to_buddhist_year(year: int) -> int:
    """Convert Christian year to Buddhist year."""
    return year + 543


def _get_thai_month_name(month: int) -> str:
    """Get Thai month name."""
    thai_months = [
        "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
        "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
    ]
    return thai_months[month - 1]


def _get_english_month_name(month: int) -> str:
    """Get English month name."""
    english_months = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    return english_months[month - 1]


class StatementPDFGenerator:
    """Generate PDF statements with Thai/English bilingual support."""
    
    def __init__(self, settings: Settings):
        self.settings = settings
        self._setup_fonts()
    
    def _setup_fonts(self):
        """Setup fonts with Thai support."""
        # Try to register Thai font if available
        try:
            thai_font_path = os.path.join(os.path.dirname(__file__), "../../assets/fonts/THSarabun.ttf")
            if os.path.exists(thai_font_path):
                pdfmetrics.registerFont(TTFont('THSarabun', thai_font_path))
                pdfmetrics.registerFont(TTFont('THSarabun-Bold', thai_font_path))
                addMapping('THSarabun', 0, 0, 'THSarabun')
                addMapping('THSarabun', 1, 0, 'THSarabun-Bold')
                self.thai_font = 'THSarabun'
            else:
                # Fallback to Helvetica (may not display Thai correctly)
                self.thai_font = 'Helvetica'
        except:
            # Fallback font
            self.thai_font = 'Helvetica'
    
    def generate_statement_pdf(self, statement: Dict[str, Any]) -> bytes:
        """Generate PDF statement from statement data."""
        buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=20*mm,
            bottomMargin=20*mm,
            leftMargin=20*mm,
            rightMargin=20*mm
        )
        
        # Build content
        elements = []
        
        # Header
        self._add_header(elements, statement)
        elements.append(Spacer(1, 10*mm))
        
        # Key amount box
        self._add_key_amount_box(elements, statement)
        elements.append(Spacer(1, 8*mm))
        
        # Summary table
        self._add_summary_table(elements, statement)
        elements.append(Spacer(1, 8*mm))
        
        # Transaction timeline
        self._add_transaction_timeline(elements, statement)
        elements.append(Spacer(1, 10*mm))
        
        # Footer disclaimers
        self._add_footer_disclaimers(elements, statement)
        
        # Build PDF
        doc.build(elements)
        
        # Return bytes
        buffer.seek(0)
        return buffer.getvalue()
    
    def _add_header(self, elements: List, statement: Dict[str, Any]):
        """Add bilingual header."""
        styles = getSampleStyleSheet()
        
        # Title Thai
        title_th_style = ParagraphStyle(
            'TitleTH',
            parent=styles['Title'],
            fontName=self.thai_font,
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=5
        )
        elements.append(Paragraph("ใบแจ้งยอดบัญชีค่าส่วนกลาง (สิ้นเดือน)", title_th_style))
        
        # Title English
        title_en_style = ParagraphStyle(
            'TitleEN',
            parent=styles['Title'],
            fontSize=14,
            alignment=TA_CENTER,
            spaceAfter=10
        )
        elements.append(Paragraph("Common Area Fee Account Statement (Month-end)", title_en_style))
        
        # Project name
        project_style = ParagraphStyle(
            'Project',
            parent=styles['Normal'],
            fontName=self.thai_font,
            fontSize=12,
            alignment=TA_CENTER,
            spaceAfter=15
        )
        elements.append(Paragraph(f"{self.settings.PROJECT_NAME_TH} / {self.settings.PROJECT_NAME_EN}", project_style))
        
        # Document info table
        header = statement['header']
        doc_id = f"STAT-{header['house_code']}-{header['period'].replace('-', '')}-{_get_short_hash(str(header))}"
        
        doc_data = [
            ["Period / ระยะเวลา:", f"{header.get('period_en', '')} / {header.get('period_th', '')}"],
            ["House / บ้านเลขที่:", header['house_code']],
            ["Owner / เจ้าของ:", header['owner_name']],
            ["Document ID / เลขที่เอกสาร:", doc_id],
            ["Generated / สร้างเอกสาร:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        doc_table = Table(doc_data, colWidths=[60*mm, 90*mm])
        doc_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.thai_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(doc_table)
    
    def _add_key_amount_box(self, elements: List, statement: Dict[str, Any]):
        """Add highlighted closing balance box."""
        closing_balance = statement['header']['closing_balance']
        
        balance_data = [[
            "ยอดคงค้างปลายเดือน / Closing Balance",
            f"THB {_format_currency(closing_balance)}"
        ]]
        
        balance_table = Table(balance_data, colWidths=[120*mm, 50*mm])
        balance_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.thai_font),
            ('FONTSIZE', (0, 0), (-1, -1), 14),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(balance_table)
    
    def _add_summary_table(self, elements: List, statement: Dict[str, Any]):
        """Add bilingual summary table."""
        summary = statement['summary']
        
        # Table headers
        summary_data = [
            ["รายการ / Description", "ภาษาอังกฤษ / English", "จำนวนเงิน / Amount (THB)"]
        ]
        
        # Summary rows
        summary_rows = [
            (summary['opening_balance']['th'], summary['opening_balance']['en'], summary['opening_balance']['amount']),
            (summary['invoices']['th'], summary['invoices']['en'], summary['invoices']['amount']),
            (summary['payments']['th'], summary['payments']['en'], summary['payments']['amount']),
            (summary['credit_notes']['th'], summary['credit_notes']['en'], summary['credit_notes']['amount']),
            (summary['closing_balance']['th'], summary['closing_balance']['en'], summary['closing_balance']['amount']),
        ]
        
        for th_text, en_text, amount in summary_rows:
            summary_data.append([
                th_text,
                en_text,
                _format_currency(amount)
            ])
        
        summary_table = Table(summary_data, colWidths=[70*mm, 70*mm, 30*mm])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.thai_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(summary_table)
    
    def _add_transaction_timeline(self, elements: List, statement: Dict[str, Any]):
        """Add transaction timeline table."""
        transactions = statement.get('transactions', [])
        
        if not transactions:
            return
        
        # Timeline header
        timeline_style = ParagraphStyle(
            'TimelineHeader',
            fontName=self.thai_font,
            fontSize=12,
            spaceAfter=5
        )
        elements.append(Paragraph("รายละเอียดรายการ / Transaction Details", timeline_style))
        
        # Table headers
        timeline_data = [
            ["วันที่ / Date", "ประเภท / Type", "อ้างอิง / Reference", "จำนวน / Amount", "คงเหลือ / Balance"]
        ]
        
        # Transaction rows
        for txn in transactions:
            timeline_data.append([
                txn['date'],
                f"{txn['type_th']} / {txn['type_en']}",
                txn['reference'],
                f"{'+' if txn['is_debit'] else '-'}{_format_currency(abs(txn['amount']))}",
                _format_currency(txn['running_balance'])
            ])
        
        timeline_table = Table(timeline_data, colWidths=[25*mm, 45*mm, 35*mm, 30*mm, 35*mm])
        timeline_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.thai_font),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (2, -1), 'LEFT'),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(timeline_table)
    
    def _add_footer_disclaimers(self, elements: List, statement: Dict[str, Any]):
        """Add footer disclaimers in Thai and English."""
        disclaimer_style = ParagraphStyle(
            'Disclaimer',
            fontName=self.thai_font,
            fontSize=10,
            alignment=TA_LEFT,
            spaceAfter=3
        )
        
        # Thai disclaimers
        thai_disclaimers = [
            "เอกสารฉบับนี้เป็นใบแจ้งยอดบัญชี มิใช่ใบเสร็จรับเงิน",
            "ยอดคงค้างคำนวณจากข้อมูล ณ สิ้นเดือนที่ระบุ",
            "หากมีข้อสงสัย กรุณาติดต่อฝ่ายบัญชี/นิติบุคคล"
        ]
        
        # English disclaimers
        english_disclaimers = [
            "This document is an account statement and is not an official receipt.",
            "Amounts are calculated as of the stated month-end.",
            "For inquiries, please contact the accounting/management office."
        ]
        
        for disclaimer in thai_disclaimers:
            elements.append(Paragraph(f"• {disclaimer}", disclaimer_style))
        
        elements.append(Spacer(1, 3))
        
        for disclaimer in english_disclaimers:
            elements.append(Paragraph(f"• {disclaimer}", disclaimer_style))
        
        # Contact info
        elements.append(Spacer(1, 5))
        contact_style = ParagraphStyle(
            'Contact',
            fontName=self.thai_font,
            fontSize=9,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(self.settings.ACCOUNTING_CONTACT, contact_style))
        
        # Page numbering
        page_style = ParagraphStyle(
            'Page',
            fontSize=9,
            alignment=TA_CENTER
        )
        elements.append(Spacer(1, 3))
        elements.append(Paragraph("หน้า 1 / 1 | Page 1 / 1", page_style))


class StatementExcelGenerator:
    """Generate Excel statements in accounting-friendly format."""
    
    def __init__(self, settings: Settings):
        self.settings = settings
    
    def generate_statement_excel(self, statement: Dict[str, Any]) -> bytes:
        """Generate Excel statement from statement data."""
        wb = Workbook()
        
        # Remove default sheet and create custom sheets
        wb.remove(wb.active)
        
        # Sheet 1: Statement
        self._create_statement_sheet(wb, statement)
        
        # Sheet 2: Raw Data
        self._create_raw_data_sheet(wb, statement)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_statement_sheet(self, wb: Workbook, statement: Dict[str, Any]):
        """Create main statement sheet."""
        ws = wb.create_sheet("Statement", 0)
        
        # Header styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        normal_font = Font(size=11)
        currency_font = Font(size=11, color="000080")
        
        # Header alignment
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        row = 1
        ws[f'A{row}'] = "ใบแจ้งยอดบัญชีค่าส่วนกลาง / Common Area Fee Account Statement"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = center_align
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # Project name
        ws[f'A{row}'] = f"{self.settings.PROJECT_NAME_TH} / {self.settings.PROJECT_NAME_EN}"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].alignment = center_align
        ws.merge_cells(f'A{row}:F{row}')
        row += 2
        
        # Document information
        header = statement['header']
        doc_info = [
            ["Period / ระยะเวลา:", f"{header.get('period_en', '')} / {header.get('period_th', '')}"],
            ["House / บ้านเลขที่:", header['house_code']],
            ["Owner / เจ้าของ:", header['owner_name']],
            ["Status / สถานะ:", header['house_status']],
            ["Generated / สร้างเวลา:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for label, value in doc_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = normal_font
            row += 1
        
        row += 1
        
        # Closing balance highlight
        ws[f'A{row}'] = "ยอดคงค้างปลายเดือน / Closing Balance:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'D{row}'] = header['closing_balance']
        ws[f'D{row}'].font = Font(bold=True, size=14, color="FF0000")
        ws[f'D{row}'].alignment = right_align
        ws[f'D{row}'].number_format = '#,##0.00'
        row += 2
        
        # Summary table
        ws[f'A{row}'] = "Summary / สรุป"
        ws[f'A{row}'].font = subheader_font
        row += 1
        
        # Summary headers
        summary_headers = ["รายการ (Thai)", "Description (English)", "Amount (THB)"]
        for col, header_text in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col, value=header_text)
            cell.font = Font(bold=True)
            cell.border = thin_border
        row += 1
        
        # Summary data
        summary = statement['summary']
        summary_rows = [
            (summary['opening_balance']['th'], summary['opening_balance']['en'], summary['opening_balance']['amount']),
            (summary['invoices']['th'], summary['invoices']['en'], summary['invoices']['amount']),
            (summary['payments']['th'], summary['payments']['en'], summary['payments']['amount']),
            (summary['credit_notes']['th'], summary['credit_notes']['en'], summary['credit_notes']['amount']),
            (summary['closing_balance']['th'], summary['closing_balance']['en'], summary['closing_balance']['amount']),
        ]
        
        for th_text, en_text, amount in summary_rows:
            ws.cell(row=row, column=1, value=th_text).border = thin_border
            ws.cell(row=row, column=2, value=en_text).border = thin_border
            amount_cell = ws.cell(row=row, column=3, value=float(amount))
            amount_cell.border = thin_border
            amount_cell.number_format = '#,##0.00'
            amount_cell.alignment = right_align
            if th_text == summary['closing_balance']['th']:
                amount_cell.font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Transaction timeline
        transactions = statement.get('transactions', [])
        if transactions:
            ws[f'A{row}'] = "Transaction Details / รายละเอียดรายการ"
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            # Transaction headers
            txn_headers = ["Date", "Type (TH)", "Type (EN)", "Reference", "Amount", "Running Balance"]
            for col, header_text in enumerate(txn_headers, 1):
                cell = ws.cell(row=row, column=col, value=header_text)
                cell.font = Font(bold=True)
                cell.border = thin_border
            row += 1
            
            # Transaction data
            for txn in transactions:
                ws.cell(row=row, column=1, value=txn['date']).border = thin_border
                ws.cell(row=row, column=2, value=txn['type_th']).border = thin_border
                ws.cell(row=row, column=3, value=txn['type_en']).border = thin_border
                ws.cell(row=row, column=4, value=txn['reference']).border = thin_border
                
                amount_cell = ws.cell(row=row, column=5, value=float(txn['amount']))
                amount_cell.border = thin_border
                amount_cell.number_format = '#,##0.00'
                amount_cell.alignment = right_align
                
                balance_cell = ws.cell(row=row, column=6, value=float(txn['running_balance']))
                balance_cell.border = thin_border
                balance_cell.number_format = '#,##0.00'
                balance_cell.alignment = right_align
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_raw_data_sheet(self, wb: Workbook, statement: Dict[str, Any]):
        """Create raw data sheet for auditing."""
        ws = wb.create_sheet("RawData", 1)
        
        # Headers
        headers = ["Date", "Type", "Reference", "Amount", "Running Balance", "Source ID"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Raw transaction data
        transactions = statement.get('transactions', [])
        for row, txn in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=txn['date'])
            ws.cell(row=row, column=2, value=txn['type_en'])
            ws.cell(row=row, column=3, value=txn['reference'])
            ws.cell(row=row, column=4, value=float(txn['amount']))
            ws.cell(row=row, column=5, value=float(txn['running_balance']))
            ws.cell(row=row, column=6, value=txn.get('source_id', 'N/A'))
        
        # Auto-adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15